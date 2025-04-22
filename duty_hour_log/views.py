# duty_hour_log/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import DutyHourLog, Initial
from .forms import DutyHourLogForm, ReportFilterForm 
from datetime import date, timedelta, datetime
from django.core.paginator import Paginator
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET  # To ensure POST for save
from django.template.loader import render_to_string # To render the static row
import json # To parse request body
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_time
from django.db.models import F, ExpressionWrapper, DurationField, Sum, Q
from django.utils import timezone
from datetime import datetime
import openpyxl # For Excel export
from .utils import format_timedelta_to_hhmm
from django.urls import reverse


# List View (Read - All) - Modified for Inline Add
def log_list(request):
    # GET Request only for displaying the list and the add button
    log_list_all = DutyHourLog.objects.all().order_by('-date', '-start_time')
    paginator = Paginator(log_list_all, 10) # Show 10 logs per page.
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Prepare data needed for the JS template row
    active_initials = list(Initial.objects.filter(is_active=True).values('id', 'code'))
    op_choices = [{'value': c[0], 'display': c[1]} for c in DutyHourLog.OP_CHOICES]
    rating_choices = [{'value': c[0], 'display': c[1]} for c in DutyHourLog.RATING_CHOICES]

    context = {
        'page_obj': page_obj,
        'active_initials_json': json.dumps(active_initials), # Pass initials as JSON
        'op_choices_json': json.dumps(op_choices), # Pass choices as JSON
        'rating_choices_json': json.dumps(rating_choices), # Pass choices as JSON
        'today_date': date.today().isoformat(), # Default date for new row
    }
    return render(request, 'duty_hour_log/log_list_inline.html', context) # Use a new template name

# New View to Save Inline Log Entry via AJAX
@require_POST # Ensures this view only accepts POST requests
def save_inline_log(request):
    try:
        data = json.loads(request.body)
        print("--- save_inline_log RECEIVED DATA ---") # <<< ADD
        print(data)                      
    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'errors': {'__all__': ['Invalid request format.']}}, status=400)

    # Instantiate the form with the received data
    form = DutyHourLogForm(data)

    if form.is_valid():
        print("--- save_inline_log FORM IS VALID ---") # <<< ADD
        try:
            new_log = form.save()
            print(f"--- save_inline_log SAVED LOG REMARKS: '{new_log.remarks}' ---") #
            # Render the static row HTML to send back to the client
            row_html = render_to_string(
                'duty_hour_log/partials/log_table_row_static.html', # A new partial template
                {'log': new_log}
            )
            return JsonResponse({
                'status': 'success',
                'new_row_html': row_html,
                'log_id': new_log.pk # Send back the ID if needed
            })
        except Exception as e:
            # Catch potential database errors during save
            print(f"--- save_inline_log SAVE EXCEPTION: {e} ---") # <<< ADD
            return JsonResponse({'status': 'error', 'errors': {'__all__': [f'Error saving log: {e}']}}, status=500)
    else:
        # Form is invalid, return errors
        print("--- save_inline_log FORM IS INVALID ---") # <<< ADD
        print(form.errors.as_json())                   # <<< ADD
        return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)


# Detail View (Read - One) - NO CHANGES NEEDED
def log_detail(request, pk):
    log = get_object_or_404(DutyHourLog, pk=pk)
    # Link to the detail page is less critical if editing happens elsewhere, but keep it for now
    # Maybe add links from the static table row later if needed
    context = {'log': log}
    return render(request, 'duty_hour_log/log_detail.html', context)


# Update View - Keep for editing existing records via separate page
def log_update(request, pk):
    log_instance = get_object_or_404(DutyHourLog, pk=pk)
    if request.method == 'POST':
        form = DutyHourLogForm(request.POST, instance=log_instance)
        if form.is_valid():
            form.save()
            # Redirect back to list maybe? Or detail?
            return redirect('log_list') # Changed redirect target
    else: # GET request
        form = DutyHourLogForm(instance=log_instance)

    context = {'form': form, 'log_instance': log_instance, 'form_title': 'Edit Duty Log'}
    # Still uses the old log_form.html - this is for editing *existing* records
    return render(request, 'duty_hour_log/log_form.html', context)

# Delete View - Keep for deleting existing records via separate page/confirmation
def log_delete(request, pk):
    log_instance = get_object_or_404(DutyHourLog, pk=pk)
    if request.method == 'POST':
        log_instance.delete()
        return redirect('log_list')

    context = {'log': log_instance}
    return render(request, 'duty_hour_log/log_confirm_delete.html', context)

@require_POST # Or use require_http_methods(['POST']) or handle PATCH
def update_inline_log(request, pk):
    try:
        log_instance = DutyHourLog.objects.get(pk=pk)
    except DutyHourLog.DoesNotExist:
        return JsonResponse({'status': 'error', 'errors': {'__all__': ['Log entry not found.']}}, status=404)

    try:
        data = json.loads(request.body)
        new_finish_time_str = data.get('finish_time', '').strip()
        new_remarks = data.get('remarks', log_instance.remarks or '').strip()

        # --- Validation ---
        errors = {}
        new_finish_time = None

        if new_finish_time_str: # Only validate if a value was provided
            new_finish_time = parse_time(new_finish_time_str)
            if new_finish_time is None:
                errors['finish_time'] = ['Invalid time format. Use HH:MM.']
            elif log_instance.start_time and new_finish_time <= log_instance.start_time:
                 errors['finish_time'] = ['Finish time must be after start time.']
        # else: finish time is being cleared or left empty, which is allowed

        if errors:
             return JsonResponse({'status': 'error', 'errors': errors}, status=400)

        # --- Update Instance ---
        log_instance.finish_time = new_finish_time # Assign None if cleared or invalid (though invalid caught above)
        log_instance.remarks = new_remarks
        log_instance.save(update_fields=['finish_time', 'remarks'])

        # --- Return Success Response ---
        return JsonResponse({
            'status': 'success',
            'log_id': log_instance.pk,
            'finish_time_display': log_instance.finish_time.strftime('%H:%M') if log_instance.finish_time else '...',
            'finish_time_value': log_instance.finish_time.strftime('%H:%M') if log_instance.finish_time else '',
            'remarks_display': log_instance.remarks[:30] + ('...' if len(log_instance.remarks) > 30 else ''), # Truncated for display
            'remarks_value': log_instance.remarks, # Full value for data attribute
            'can_edit_again': not bool(log_instance.finish_time) # True if finish_time is None/empty
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'errors': {'__all__': ['Invalid request format.']}}, status=400)
    except Exception as e:
        # Log the exception e
        print(f"Error updating inline log {pk}: {e}") # Basic logging
        return JsonResponse({'status': 'error', 'errors': {'__all__': [f'An unexpected server error occurred.']}}, status=500)

def report_view(request):
    form = ReportFilterForm(request.GET or None) # Populate with GET data if submitted
    total_duration_timedelta = None
    total_hours = 0
    filtered_logs = DutyHourLog.objects.none() # Start with an empty queryset
    submitted_filters = {} # To display applied filters

    if form.is_valid():
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        initial = form.cleaned_data['initial']
        op = form.cleaned_data['op']
        rating = form.cleaned_data['rating']

        # Store submitted filters for display
        submitted_filters = {
            'start_date': start_date.isoformat(), # Correct
            'end_date': end_date.isoformat(),     # Correct
            'initial': initial.id if initial else '', # PASS ID OR EMPTY STRING - Correct
            'initial_code': initial.code if initial else 'Any', # For display
            'op': op if op else '', # PASS VALUE OR EMPTY STRING - Correct
            'op_display': dict(DutyHourLog.OP_CHOICES).get(op, 'Any') if op else 'Any', # For display
            'rating': rating if rating else '', # PASS VALUE OR EMPTY STRING - Correct
            'rating_display': dict(DutyHourLog.RATING_CHOICES).get(rating, 'Any') if rating else 'Any', # For display
        }

        # Base query: within date range and MUST have finish_time
        logs_query = DutyHourLog.objects.filter(
            date__range=[start_date, end_date],
            finish_time__isnull=False # Important: Only include logs with finish times
        )

        # Apply optional filters
        if initial:
            # Filter logs where the selected initial was involved in *any* relevant role
            logs_query = logs_query.filter(
                Q(initial=initial) | Q(trainee=initial) | Q(ojti=initial) | Q(examiner=initial)
            )
        if op:
            logs_query = logs_query.filter(op=op)
        if rating:
            logs_query = logs_query.filter(rating=rating)

        # Annotate with duration and calculate total duration
        # Ensure calculation handles time fields correctly (database might store as time without date)
        # This assumes start_time and finish_time are on the *same day* as specified by 'date'
        # NOTE: SQLite has limitations with DurationField arithmetic. This might work better on PostgreSQL/MySQL.
        # If issues arise on SQLite, consider calculating in Python (less efficient for large datasets).
        try:
             annotated_logs = logs_query.annotate(
                duration=ExpressionWrapper(
                    F('finish_time') - F('start_time'),
                    output_field=DurationField()
                 )
             )
             # Handle potential negative durations if finish_time is on the next day (unlikely but possible)
             # For simplicity, we assume finish_time > start_time on the same day.
             # A more robust solution would involve datetime objects.

             # Aggregate the sum of durations
             aggregation_result = annotated_logs.aggregate(total_duration=Sum('duration'))
             total_duration_timedelta = aggregation_result.get('total_duration')

        except Exception as e:
            # Catch potential database-specific errors (especially with SQLite time subtraction)
            print(f"Database aggregation error: {e}. Falling back to Python calculation.")
            total_duration_timedelta = timedelta(0)
            logs_for_python_calc = list(logs_query) # Evaluate the query
            for log in logs_for_python_calc:
                # Create datetime objects for calculation if needed (assuming same day)
                # This basic approach works if start/finish are TimeFields on same date
                 if log.start_time and log.finish_time and log.finish_time > log.start_time:
                    # Simplistic calculation for TimeFields (might be inaccurate across midnight)
                    # A robust way requires combining DateField and TimeField into DateTimeFields
                     start_dt = timezone.make_aware(datetime.combine(log.date, log.start_time))
                     finish_dt = timezone.make_aware(datetime.combine(log.date, log.finish_time))
                     total_duration_timedelta += (finish_dt - start_dt)
                 elif log.start_time and log.finish_time and log.finish_time < log.start_time:
                     # Handle overnight scenario simply (adds 24h) - adjust if needed
                     start_dt = timezone.make_aware(datetime.combine(log.date, log.start_time))
                     # Assume finish is on the *next* day relative to start
                     finish_dt = timezone.make_aware(datetime.combine(log.date + timedelta(days=1), log.finish_time))
                     total_duration_timedelta += (finish_dt - start_dt)


        if total_duration_timedelta:
            total_seconds = total_duration_timedelta.total_seconds()
            total_hours = total_seconds / 3600  # Convert seconds to hours
            filtered_logs = logs_query # Assign the filtered logs for potential display

    context = {
        'form': form,
        'total_hours': total_hours if total_duration_timedelta is not None else None,
        'submitted_filters': submitted_filters if form.is_valid() else None,
        'filtered_logs': filtered_logs, # Pass logs for optional display
        'has_results': total_duration_timedelta is not None and form.is_valid(),
    }
    return render(request, 'duty_hour_log/report.html', context)

@require_GET # Only allow GET requests for export
def export_report_to_excel(request):
    print(f"--- export_report_to_excel called ---") # DEBUG
    print(f"Request GET data: {request.GET}")      # DEBUG

    # Explicitly use request.GET, None is less common here
    form = ReportFilterForm(request.GET)

    if form.is_valid():
        print("Form IS valid. Proceeding with Excel generation.") # DEBUG
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        initial = form.cleaned_data['initial']
        op = form.cleaned_data['op']
        rating = form.cleaned_data['rating']

        # --- Re-apply the *exact same* filtering logic ---
        # ... (filtering logic remains the same) ...
        logs_query = DutyHourLog.objects.filter(
            date__range=[start_date, end_date],
            finish_time__isnull=False
        ).select_related('initial', 'trainee', 'ojti', 'examiner').order_by('date', 'start_time')

        if initial:
            logs_query = logs_query.filter(
                Q(initial=initial) | Q(trainee=initial) | Q(ojti=initial) | Q(examiner=initial)
            )
        if op:
            logs_query = logs_query.filter(op=op)
        if rating:
            logs_query = logs_query.filter(rating=rating)

        # ... (Excel generation logic remains the same) ...
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Duty Log Report"
        headers = [
            "Date", "Start Time", "Finish Time", "Duration (HH:MM)",
            "Operation Type", "Rating", "Initial (Solo)", "Trainee", "OJTI", "Examiner",
            "Remarks"
        ]
        ws.append(headers)
        for log in logs_query:
            duration = log.calculate_duration()
            duration_str = format_timedelta_to_hhmm(duration) if duration else "-"
            row_data = [
                log.date, log.start_time, log.finish_time, duration_str,
                log.get_op_display(), log.get_rating_display(),
                log.initial.code if log.initial else "-",
                log.trainee.code if log.trainee else "-",
                log.ojti.code if log.ojti else "-",
                log.examiner.code if log.examiner else "-",
                log.remarks if log.remarks else ""
            ]
            ws.append(row_data)
            ws.cell(row=ws.max_row, column=1).number_format = 'YYYY-MM-DD'
            ws.cell(row=ws.max_row, column=2).number_format = 'HH:MM'
            ws.cell(row=ws.max_row, column=3).number_format = 'HH:MM'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"duty_log_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        print("Excel response prepared and returned.") # DEBUG
        return response

    else:
        # Form is invalid
        print(f"Form IS INVALID.") # DEBUG
        print(f"Form errors: {form.errors.as_json()}") # DEBUG: Keep this for detailed JSON errors

        # Construct a simpler error message for display
        error_details = []
        for field, errors in form.errors.items():
            # errors is likely an ErrorList, join its messages
            error_details.append(f"{field.replace('_',' ').title()}: {'. '.join(errors)}")

        error_msg = f"Invalid filters provided for export. Please check report filters. Details: {'; '.join(error_details)}"
        messages.error(request, error_msg)

        # Redirect back to report page, maybe preserving original params
        original_query_string = request.GET.urlencode()
        redirect_url = f"{reverse('report_view')}?{original_query_string}"
        print(f"Redirecting to: {redirect_url}") # DEBUG
        return redirect(redirect_url)